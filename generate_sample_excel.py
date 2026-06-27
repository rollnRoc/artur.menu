import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_sample_excel():
    base_dir = r"C:\Users\Emre\.gemini\antigravity\scratch\artur_siparis"
    dest_path = os.path.join(base_dir, "admin", "Ornek_Menu.xlsx")
    
    # Create a new workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Menü Listesi"
    
    # Set headers
    headers = ["Kategori", "Ürün Adı", "Ürün Fiyatı", "Resim Linki"]
    ws.append(headers)
    
    # Add sample data
    sample_data = [
        ["FASTFOOD", "Cheeseburger", 250, "https://images.unsplash.com/photo-1568901346375-23c9450c58cd?w=500"],
        ["FASTFOOD", "Klasik Hamburger", 220, "https://images.unsplash.com/photo-1586190848861-99aa4a171e90?w=500"],
        ["FASTFOOD", "Patates Kızartması", 90, ""],
        ["İÇECEKLER", "Kutu Kola", 50, ""],
        ["İÇECEKLER", "Taze Sıkılmış Portakal Suyu", 80, ""],
        ["TATLILAR", "Çikolatalı Sufle", 120, "https://images.unsplash.com/photo-1606313564200-e75d5e30476c?w=500"],
        ["TATLILAR", "Dondurmalı İrmik Helvası", 110, ""]
    ]
    
    for row in sample_data:
        ws.append(row)
        
    # Styles
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, bold=False)
    
    # Dark blue fill for header: #1F4E78
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Style Header Row
    for col_idx in range(1, 5):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center if col_idx != 2 else align_left
        cell.border = thin_border
        
    # Style Data Rows
    for row_idx in range(2, len(sample_data) + 2):
        # Kategori
        cell = ws.cell(row=row_idx, column=1)
        cell.font = data_font
        cell.alignment = align_center
        cell.border = thin_border
        
        # Ürün Adı
        cell = ws.cell(row=row_idx, column=2)
        cell.font = data_font
        cell.alignment = align_left
        cell.border = thin_border
        
        # Ürün Fiyatı
        cell = ws.cell(row=row_idx, column=3)
        cell.font = data_font
        cell.alignment = align_right
        cell.number_format = "#,##0.00"
        cell.border = thin_border
        
        # Resim Linki
        cell = ws.cell(row=row_idx, column=4)
        cell.font = data_font
        cell.alignment = align_left
        cell.border = thin_border

    # Adjust column widths
    column_widths = {
        'A': 18,  # Kategori
        'B': 30,  # Ürün Adı
        'C': 15,  # Ürün Fiyatı
        'D': 50   # Resim Linki
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    # Make sure output directory exists
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    
    wb.save(dest_path)
    print(f"Sample Excel Menu saved to: {dest_path}")
    return True

if __name__ == "__main__":
    create_sample_excel()
