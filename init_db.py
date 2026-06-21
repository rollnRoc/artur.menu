import openpyxl
import json
import os

db_dir = os.path.join("C:\\Users\\Emre\\.gemini\\antigravity\\scratch\\artur_siparis", "data")
os.makedirs(db_dir, exist_ok=True)

# 1. Read images from Dışardan Sipariş Alma.xlsx
images_map = {}
orig_file = "C:\\Users\\Emre\\.gemini\\antigravity\\scratch\\Dışardan Sipariş Alma.xlsx"
if os.path.exists(orig_file):
    print("Reading image links...")
    wb_orig = openpyxl.load_workbook(orig_file)
    if "Menü" in wb_orig.sheetnames:
        sheet = wb_orig["Menü"]
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(h).strip().lower() for h in rows[0]]
        name_idx = headers.index("ürün adı") if "ürün adı" in headers else 1
        img_idx = headers.index("image") if "image" in headers else 3
        
        for row in rows[1:]:
            if len(row) > max(name_idx, img_idx):
                name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
                img = str(row[img_idx]).strip() if row[img_idx] is not None else ""
                if name and img and img.startswith("http"):
                    images_map[name] = img
    print(f"Loaded image links.")

# 2. Read products from Ürünler.xlsx
menu_items = []
prod_file = "C:\\Users\\Emre\\.gemini\\antigravity\\scratch\\Ürünler.xlsx"
if os.path.exists(prod_file):
    print("Reading products...")
    wb_prod = openpyxl.load_workbook(prod_file)
    sheet = wb_prod.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(h).strip().lower() for h in rows[0]]
    
    cat_idx = headers.index("kategori") if "kategori" in headers else 0
    name_idx = headers.index("ürün adı") if "ürün adı" in headers else 1
    price_idx = headers.index("ürün fiyatı") if "ürün fiyatı" in headers else 2
    
    item_id = 1
    for row in rows[1:]:
        if len(row) > max(cat_idx, name_idx, price_idx) and row[name_idx] is not None:
            cat = str(row[cat_idx]).strip()
            name = str(row[name_idx]).strip()
            try:
                price = float(row[price_idx]) if row[price_idx] is not None else 0.0
                if price.is_integer():
                    price = int(price)
            except ValueError:
                price = 0
                
            img = images_map.get(name, "")
            
            menu_items.append({
                "id": item_id,
                "category": cat,
                "name": name,
                "price": price,
                "image": img,
                "active": True
            })
            item_id += 1
            
# 3. Create database.json
database = {
    "menu": menu_items,
    "orders": [],
    "settings": {
        "orderCounter": 0,
        "lastCounterDate": ""
    }
}

db_file = os.path.join(db_dir, "database.json")
with open(db_file, "w", encoding="utf-8") as f:
    json.dump(database, f, indent=2, ensure_ascii=False)

print(f"Successfully created database.json")
